import json
import logging
import os
import pandas as pd 
import psycopg2
import hashlib
from datetime import datetime, date
from openpyxl.styles import Font
from openpyxl.comments import Comment

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

# Variables globales que se inicializarán después de leer los datos
HUB = None
PROVEEDOR = None
HOSTNAME = None
TIPO_DE_RED = None
TIPO_DE_ZONA = None
TIPO_DE_COBERTURA = None
REGION = None
ZONA = None
PARROQUIA = None
FEEDER = None
CLUSTER = None
HORIZONTAL_RESIDENCIAL_HPs = None
HORIZONTAL_COMERCIAL_HPs = None
VERTICAL_RESIDENCIAL_HPs = None
VERTICAL_COMERCIAL_HPs = None
CANTIDAD_EDIFICIOS_PROYECTADOS = None
EDIF_RESID_PROYECTADOS_HPs = None
EDIF_COMERCIAL_PROYECTADOS_HPs = None
SOLARES = None
HPs_TOTALES = None
PUERTOS_HABILITADOS = None
HOME_PASSES_TOTAL = None
BUSINESS_PASSES_TOTAL = None

def conexion_bd():
    try:
        with open("configuracion/conexion.json", "r") as archivo:
            credenciales = json.load(archivo)
        
        conn = psycopg2.connect(
            dbname=credenciales["PostgresSQL"]["database"],
            user=credenciales["PostgresSQL"]["user"],
            password=credenciales["PostgresSQL"]["password"],
            host=credenciales["PostgresSQL"]["host"],
            port=credenciales["PostgresSQL"]["port"]
        )
        
        if conn is not None:
            logging.info("Conexión exitosa a la base de datos")
            return conn
    except Exception as e:
        logging.error(f"Error al conectar a la base de datos: {e}")
        return None

def lectura_data():
    try:
        df = pd.read_excel("data/data.xlsx", sheet_name="Liberacion")
        if df.empty:
            logging.error("El archivo no contiene datos")
            return None
        else:
            logging.info("Datos leídos correctamente")
            #obtenemos las variables
            hub = df["HUB"].iloc[0]
            proveedor=df["PROVEEDOR"].iloc[0]
            hostname = df["HOSTNAME"].iloc[0]
            tipo_red = df["TIPO DE RED"].iloc[0]
            tipo_zona = df["TIPO DE ZONA"].iloc[0]
            tipo_cobertura=df["TIPO DE COBERTURA"].iloc[0]
            region=df["REGIÓN"].iloc[0]
            zona=df["ZONA"].iloc[0]
            parroquia=df["PARROQUIA"].iloc[0]
            feeder=df["FEEDER"].iloc[0]
            cluster=df["CLUSTER"].iloc[0]
            horizontal_residencial=df["Horizontal Residencial (HPs)"].iloc[0]
            horizontal_comercial=df["Horizontal Comercial (HPs)"].iloc[0]
            vertical_residencial=df["Vertical Residencial (HPs)"].iloc[0]
            vertical_comercial=df["Vertical Comercial (HPs)"].iloc[0]
            cantidad_edficios_proyectados=df["Cantidad de Edificios Proyectados"].iloc[0]
            edificio_residencial_proyectado=df["Edif Resid Proyectados (HPs)"].iloc[0]
            edificio_comercial_proyectado=df["Edif Comercial Proyectados (HPs)"].iloc[0]
            solares=df["Solares"].iloc[0]
            hp_total=df["HP'S TOTALES"].iloc[0]
            puertos_habilitados=df["PUERTOS HABILITADOS"].iloc[0]
            #Comprobar region 
            if (region=='R1'):
                #leer Correo_R1.md
                with open("correos/Correo_R1.md", "r") as archivo:
                    correos = archivo.read()
                    print(f"\n---Correos de R1---\n{correos}\n")
            else:
                #leer Correo_R2.md
                with open("correos/Correo_R2.md", "r") as archivo:
                    correos = archivo.read()
                    print(f"\n---Correos de R2---\n{correos}\n")
            return hub, proveedor, hostname, tipo_red, tipo_zona, tipo_cobertura, region, zona, parroquia, feeder, cluster, horizontal_residencial, horizontal_comercial, vertical_residencial, vertical_comercial, cantidad_edficios_proyectados, edificio_residencial_proyectado, edificio_comercial_proyectado, solares, hp_total, puertos_habilitados
    except Exception as e:
        logging.error(f"Error al leer el archivo: {e}")
        return None

def inicializar_variables_globales():
    global HUB, PROVEEDOR, HOSTNAME, TIPO_DE_RED, TIPO_DE_ZONA, TIPO_DE_COBERTURA 
    global REGION, ZONA, PARROQUIA, FEEDER, CLUSTER
    global HORIZONTAL_RESIDENCIAL_HPs, HORIZONTAL_COMERCIAL_HPs
    global VERTICAL_RESIDENCIAL_HPs, VERTICAL_COMERCIAL_HPs
    global CANTIDAD_EDIFICIOS_PROYECTADOS, EDIF_RESID_PROYECTADOS_HPs
    global EDIF_COMERCIAL_PROYECTADOS_HPs, SOLARES, HPs_TOTALES, PUERTOS_HABILITADOS
    global HOME_PASSES_TOTAL, BUSINESS_PASSES_TOTAL
    
    resultado = lectura_data()
    if not resultado:
        logging.error("No se pudieron inicializar las variables globales")
        return False
        
    (HUB, PROVEEDOR, HOSTNAME, TIPO_DE_RED, TIPO_DE_ZONA, TIPO_DE_COBERTURA,
     REGION, ZONA, PARROQUIA, FEEDER, CLUSTER, HORIZONTAL_RESIDENCIAL_HPs,
     HORIZONTAL_COMERCIAL_HPs, VERTICAL_RESIDENCIAL_HPs, VERTICAL_COMERCIAL_HPs,
     CANTIDAD_EDIFICIOS_PROYECTADOS, EDIF_RESID_PROYECTADOS_HPs,
     EDIF_COMERCIAL_PROYECTADOS_HPs, SOLARES, HPs_TOTALES, PUERTOS_HABILITADOS) = resultado
    
    # Calcular los totales de home passes y business passes
    HOME_PASSES_TOTAL = HPs_TOTALES - HORIZONTAL_COMERCIAL_HPs - VERTICAL_COMERCIAL_HPs
    BUSINESS_PASSES_TOTAL = HORIZONTAL_COMERCIAL_HPs + VERTICAL_COMERCIAL_HPs
    
    return True

# Corrección de la función id_hash_cluster() en Liberacion.py
def id_hash_cluster():
    """Genera un ID único para el cluster basado en la fecha actual y el nombre del cluster."""
    try:
        fecha_hoy = datetime.now().strftime("%Y%m%d")
        cadena = f"{fecha_hoy}{CLUSTER}"
        # Usar MD5 para mantener compatibilidad con lo mencionado en los comentarios
        id_hash = hashlib.md5(cadena.encode()).hexdigest()
        return id_hash
    except Exception as e:
        logging.error(f"Error al generar hash: {e}")
        # Generar un hash aleatorio como fallback
        return hashlib.md5(str(datetime.now().timestamp()).encode()).hexdigest()
def prueba():
    if not inicializar_variables_globales():
        print("No se pudieron leer los datos")
        return
    
    # Verificar si el cluster existe en la BD
    conn = conexion_bd()
    if conn is None:
        logging.error("No se pudo conectar a la base de datos")
        return
    
    try:
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM clusters WHERE nombre = '{CLUSTER}'")
        existe_cluster = cursor.fetchone() is not None
        
        if existe_cluster:
            caso_existencia()
        else:
            caso_liberacion()
    except Exception as e:
        logging.error(f"Error al consultar el cluster: {e}")
    finally:
        if conn:
            cursor.close()
            conn.close()

#función para exportar excel
def exportar_excel_alcance(datos, ruta_archivo=None):
    """Exporta los datos de la consulta a un archivo Excel con mejor rendimiento"""
    if ruta_archivo is None:
        ruta_archivo = f"generador/alcance_{CLUSTER}.xlsx"
        
    try:
        # Asegurar que el directorio existe
        os.makedirs(os.path.dirname(ruta_archivo), exist_ok=True)

        # Definir los encabezados
        encabezados = [
            'id', 'hostname', 'nombre', 'zona_cobertura', 'canton',
            'puertos_habilitados', 'hps_liberadas', 'home_passes',
            'business_passes', 'fecha_liberacion', 'hp_horizontal_res',
            'hp_horizontal_com', 'hp_vertical_res', 'hp_vertical_com',
            'edif_res', 'edif_com', 'solares_res', 'tipo_cobertura',
            'region', 'parroquia', 'observacion', 'tipo_red',
            'fecha_liberacion_corp', 'tipo', 'tipo_zona'
        ]

        # Crear DataFrame directamente con los encabezados correctos para evitar reindexación
        df = pd.DataFrame(datos, columns=encabezados)
        
        # Calcular eficientemente totales numéricos usando vectorización pandas
        campos_numericos = [
            'puertos_habilitados', 'hps_liberadas', 'home_passes', 'business_passes',
            'hp_horizontal_res', 'hp_horizontal_com', 'hp_vertical_res', 'hp_vertical_com',
            'edif_res', 'edif_com', 'solares_res'
        ]
        
        # Calcular sumas en una sola operación vectorizada
        sumas = df[campos_numericos].sum()
        
        # Construir fila de diferencias entre variables globales y BD
        hoy = date.today()
        
        # Crear diccionario de totales directamente
        totales = {
            'id': id_hash_cluster(),
            'hostname': df.iloc[-1]['hostname'] if not df.empty else HOSTNAME,
            'nombre': df.iloc[-1]['nombre'] if not df.empty else CLUSTER,
            'zona_cobertura': df.iloc[-1]['zona_cobertura'] if not df.empty else ZONA,
            'canton': df.iloc[-1]['canton'] if not df.empty else 'SAMBORONDON',
            'tipo_cobertura': df.iloc[-1]['tipo_cobertura'] if not df.empty else TIPO_DE_COBERTURA,
            'region': df.iloc[-1]['region'] if not df.empty else REGION,
            'parroquia': df.iloc[-1]['parroquia'] if not df.empty else PARROQUIA,
            'tipo_red': df.iloc[-1]['tipo_red'] if not df.empty else TIPO_DE_RED,
            'tipo': df.iloc[-1]['tipo'] if not df.empty else 'N/A',
            'tipo_zona': TIPO_DE_ZONA,
            'fecha_liberacion': hoy,
            'fecha_liberacion_corp': hoy,
            'puertos_habilitados': PUERTOS_HABILITADOS - sumas['puertos_habilitados'],
            'hps_liberadas': HPs_TOTALES - sumas['hps_liberadas'],
            'home_passes': HOME_PASSES_TOTAL - sumas['home_passes'],
            'business_passes': BUSINESS_PASSES_TOTAL - sumas['business_passes'],
            'hp_horizontal_res': HORIZONTAL_RESIDENCIAL_HPs - sumas['hp_horizontal_res'],
            'hp_horizontal_com': HORIZONTAL_COMERCIAL_HPs - sumas['hp_horizontal_com'],
            'hp_vertical_res': VERTICAL_RESIDENCIAL_HPs - sumas['hp_vertical_res'],
            'hp_vertical_com': VERTICAL_COMERCIAL_HPs - sumas['hp_vertical_com'],
            'edif_res': EDIF_RESID_PROYECTADOS_HPs - sumas['edif_res'],
            'edif_com': EDIF_COMERCIAL_PROYECTADOS_HPs - sumas['edif_com'],
            'solares_res': SOLARES - sumas['solares_res'],
        }
        
        # Añadir observación
        ultimo_registro = df.iloc[-1] if not df.empty else None
        if ultimo_registro is not None:
            totales['observacion'] = f"{ultimo_registro['observacion']} - Pendiente: {totales['home_passes']} Home Passes"
        else:
            totales['observacion'] = f"Feeder: {FEEDER}, Hub: {HUB}"
        
        # Crear DataFrame de totales y concatenar eficientemente
        df_final = pd.concat([df, pd.DataFrame([totales])], ignore_index=True)
        
        # Usar context manager para asegurar cierre de recursos
        with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='Resultados')
            
            # Obtener la hoja de trabajo
            worksheet = writer.sheets['Resultados']
            
            # Formatear última fila de manera más eficiente
            ultima_fila = len(df_final) + 1
            negrita = Font(bold=True)
            
            # Aplicar formato en lote para mejor rendimiento
            for col in range(1, len(encabezados) + 1):
                worksheet.cell(row=ultima_fila, column=col).font = negrita
                
            # Añadir comentario explicativo
            comment = Comment(text="Valores pendientes: Variables globales menos los registros existentes.", author="Sistema")
            worksheet.cell(row=ultima_fila, column=1).comment = comment
            
        print(f"Archivo Excel exportado en {ruta_archivo}")
        return ruta_archivo
        
    except Exception as e:
        print(f'Error al exportar a Excel: {str(e)}')
        return None

def caso_existencia():
    conexion = conexion_bd() # Conectar a la BD
    if conexion is not None:
        try:
            cursor = conexion.cursor()
            query = f"SELECT * FROM clusters WHERE nombre = '{CLUSTER}'"
            cursor.execute(query)
            result = cursor.fetchall()
            if result:
                print(f"Se encontraron {len(result)} registros para el cluster {CLUSTER}.")
                # Exportar resultados a Excel
                archivo = exportar_excel_alcance(result)
                if archivo:
                    print(f"Consulta el archivo {archivo} para ver los resultados detallados.")
                return True
            return False
        except Exception as e:
            print(f"Error en la consulta: {str(e)}")
            return False
        finally:
            cursor.close()
            conexion.close()
    return False

def caso_liberacion():
    """
    Función para crear un archivo Excel con datos de un nuevo cluster que no existe en la BD.
    Usa las variables globales para llenar los campos y genera un ID único.
    """
    try:
        # Generar el ID único
        id_cluster = id_hash_cluster()
        
        # Fecha actual
        hoy = date.today()
        
        # Crear un registro con los datos de las variables globales
        datos = [{
            'id': id_cluster,
            'hostname': HOSTNAME,
            'nombre': CLUSTER,
            'zona_cobertura': ZONA,
            'canton': 'SAMBORONDON',
            'puertos_habilitados': PUERTOS_HABILITADOS,
            'hps_liberadas': HPs_TOTALES, 
            'home_passes': HOME_PASSES_TOTAL,
            'business_passes': BUSINESS_PASSES_TOTAL,
            'fecha_liberacion': hoy,  # Fecha actual para liberación
            'hp_horizontal_res': HORIZONTAL_RESIDENCIAL_HPs,
            'hp_horizontal_com': HORIZONTAL_COMERCIAL_HPs,
            'hp_vertical_res': VERTICAL_RESIDENCIAL_HPs,
            'hp_vertical_com': VERTICAL_COMERCIAL_HPs,
            'edif_res': EDIF_RESID_PROYECTADOS_HPs,
            'edif_com': EDIF_COMERCIAL_PROYECTADOS_HPs,
            'solares_res': SOLARES,
            'tipo_cobertura': TIPO_DE_COBERTURA,
            'region': REGION,
            'parroquia': PARROQUIA,
            'observacion': f"Feeder: {FEEDER}, Hub: {HUB}",
            'tipo_red': TIPO_DE_RED,
            'fecha_liberacion_corp': hoy,  # Fecha actual para liberación corporativa
            'tipo': 'N/A',
            'tipo_zona': TIPO_DE_ZONA
        }]
        
        # Crear DataFrame
        df = pd.DataFrame(datos)
        
        # Generar nombre de archivo
        ruta_archivo = f"generador/liberacion_{CLUSTER}.xlsx"
        
        # Asegurar que el directorio existe
        os.makedirs(os.path.dirname(ruta_archivo), exist_ok=True)
        
        # Exportar a Excel con formato
        with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Liberación')
            
            # Aplicar formato negrita a las celdas (encabezados)
            worksheet = writer.sheets['Liberación']
            for col in range(1, len(df.columns) + 1):
                celda = worksheet.cell(row=1, column=col)
                celda.font = Font(bold=True)
                
        print(f"Archivo de liberación creado exitosamente en {ruta_archivo}")
        return ruta_archivo
        
    except Exception as e:
        print(f"Error en caso_liberación: {str(e)}")
        return None

def main():
    prueba()

if __name__ == "__main__":
    main()
